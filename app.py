import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import pandas as pd
import threading
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from PIL import Image, ImageTk

# Create the main window
root = tk.Tk()
root.title("Healthcare Finance Calc App")

app_icon_path = "./images/app_icon.png"  # Ensure this PNG exists in your project folder
img = tk.PhotoImage(file=app_icon_path)
root.iconphoto(True, img)  
root.state('zoomed')  # Fullscreen mode
root.resizable(False, False)  # Disable resizing
root.configure(bg="white")  # Light mode

# Frame for layout
file_frame = tk.Frame(root, bg="white")
file_frame.pack(pady=20, padx=20)

# Load File Upload Icon
icon_upload = Image.open("./images/fileUploadIcon.png")  # Replace with actual path
icon_upload = icon_upload.resize((35,35), Image.Resampling.LANCZOS)
icon_upload_tk = ImageTk.PhotoImage(icon_upload)

# Row 1: File selection label
file_label = tk.Label(file_frame, text="Upload a CSV or Excel file", font=("Arial", 14), bg="white", fg="black")
file_label.grid(row=0, column=0, columnspan=2, pady=(0, 15))  # Centered with spacing

# Styled File Upload Button (Matches Other Buttons)
upload_button = tk.Button(
    file_frame,
    text=" Choose File",
    command=lambda: upload_file(),
    font=("Arial", 12, "bold"),
    bg="#007bff",
    padx=10, pady=10,
    bd=2, relief="raised",
    image=icon_upload_tk,
    compound="left",  # Icon on left, text on right
)
upload_button.grid(row=1, column=0, pady=(5, 20), padx=10)

# Keep reference to prevent garbage collection
upload_button.image = icon_upload_tk

# Variable to store the loaded data
data = None
output_data = []  # To store calculation results

# Treeview frame
treeview_frame = tk.Frame(root, bg="white", highlightbackground="black", highlightthickness=3)
treeview_frame.pack(pady=10, fill="both", expand=True)

# Treeview widget to display the table
treeview = ttk.Treeview(treeview_frame, show="headings")
treeview.pack(side=tk.LEFT, fill="both", expand=True)


# Configure Treeview style to show borders
style = ttk.Style()
style.configure("Treeview", background="white", foreground="black", rowheight=25, fieldbackground="white")
style.configure("Treeview.Heading", font=("Arial", 12, "bold"), background="lightgray", foreground="black")
treeview.tag_configure("oddrow", background="white")
treeview.tag_configure("evenrow", background="#f0f0f0")

# Scrollbars
vertical_scrollbar = ttk.Scrollbar(treeview_frame, orient="vertical", command=treeview.yview)
vertical_scrollbar.pack(side="right", fill="y")
treeview.configure(yscrollcommand=vertical_scrollbar.set)


horizontal_scrollbar = ttk.Scrollbar(treeview_frame, orient="horizontal", command=treeview.xview)
horizontal_scrollbar.pack(side="bottom", fill="x")
treeview.configure(xscrollcommand=horizontal_scrollbar.set)

# Button frame
import tkinter as tk
from PIL import Image, ImageTk

# Load button icons
icon_run = Image.open("./images/run_icon.png")  # Replace with actual path
icon_run = icon_run.resize((35,35), Image.Resampling.LANCZOS)
icon_run_tk = ImageTk.PhotoImage(icon_run)

icon_download = Image.open("./images/download_icon.png")  # Replace with actual path
icon_download = icon_download.resize((35,35), Image.Resampling.LANCZOS)
icon_download_tk = ImageTk.PhotoImage(icon_download)

icon_unique = Image.open("./images/unique_icon.png")  # Replace with actual path
icon_unique = icon_unique.resize((35,35), Image.Resampling.LANCZOS)
icon_unique_tk = ImageTk.PhotoImage(icon_unique)

# Create a frame for buttons
button_frame = tk.Frame(root, bg="white")
button_frame.pack(pady=10)

# Styled Run Calculation Button
run_button = tk.Button(
    button_frame,
    text=" Run Calculation",
    command=lambda: start_calculation(),
    font=("Arial", 12, "bold"),
    bg="#4CAF50",  # Green background 
    padx=10, pady=10,
    bd=2, relief="raised",
    image=icon_run_tk,  # Add icon
    compound="left",  # Icon on left, text on right
)
run_button.pack(side=tk.LEFT, padx=10, pady=5)

# Styled Download Excel Button
download_button = tk.Button(
    button_frame,
    text=" Download Excel",
    command=lambda: start_download(),
    font=("Arial", 12, "bold"),
    bg="#2196F3",  # Blue background
    padx=10, pady=10,
    bd=2, relief="raised",
    image=icon_download_tk,
    compound="left",
)
download_button.pack(side=tk.LEFT, padx=10, pady=5)

# Styled Download Unique Calculation Button
download_unique_button = tk.Button(
    button_frame,
    text=" Download Unique Calculation",
    command=lambda: start_download_unique(),
    font=("Arial", 12, "bold"),
    bg="#FF9800",  # Orange background
    padx=10, pady=10,
    bd=2, relief="raised",
    image=icon_unique_tk,
    compound="left",
)
download_unique_button.pack(side=tk.LEFT, padx=10, pady=5)

# Keep references to images to prevent garbage collection
run_button.image = icon_run_tk
download_button.image = icon_download_tk
download_unique_button.image = icon_unique_tk

# Loading Label
loading_label = tk.Label(root, text="", font=("Arial", 12, "italic"), bg="white", fg="blue")
loading_label.pack(pady=10)

# Function to show a loading message
def show_loading(message):
    loading_label.config(text=message)
    root.update_idletasks()  # Forces UI update

def hide_loading():
    loading_label.config(text="")

# Function to load and display the file preview in a table
def upload_file():
    global data
    file_path = filedialog.askopenfilename(
        title="Select a CSV or Excel file",
        filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("Excel files", "*.xls")]
    )

    if not file_path:
        print("No file selected.")
        return

    # Show loading message before starting
    show_loading("Loading file...")

    # Delay file processing to allow UI update
    root.after(100, lambda: load_file(file_path))  # Runs after 100ms

def load_file(file_path):
    global data
    try:
        print(f"Selected file: {file_path}")  # Debugging

        # Read CSV or Excel file
        if file_path.endswith('.csv'):
            data = pd.read_csv(file_path, encoding="utf-8", engine="python")
        else:
            data = pd.read_excel(file_path, engine="openpyxl")

        print(f"File loaded successfully with shape: {data.shape}")  # Debugging

        # Data Cleaning
        data.columns = data.columns.str.strip().str.lower()
        data = data.dropna(axis=1, how='all')
        data = data.fillna("--")

        print(f"Processed Columns: {data.columns.tolist()}")  # Debugging

        # Clear old data from Treeview
        for row in treeview.get_children():
            treeview.delete(row)

        # Configure Treeview
        treeview["columns"] = list(data.columns)
        treeview["show"] = "headings"

        for col in data.columns:
            treeview.heading(col, text=col, anchor="w")
            treeview.column(col, width=150, anchor="w")

        # Insert Data into Treeview
        for _, row in data.iterrows():
            treeview.insert("", "end", values=row.tolist())

        # Hide Upload Button, Show Run Button
        if "button" in globals():
            button.pack_forget()
        if "run_button" in globals():
            run_button.pack(side=tk.LEFT, padx=10)

    except Exception as e:
        print(f"❌ Error while loading file: {str(e)}")
        messagebox.showerror("Error", f"Failed to load the file.\nDetails: {str(e)}")

    finally:
        hide_loading()

# Function to start calculation in a separate thread
def start_calculation():
    threading.Thread(target=run_calculation).start()

# Function to run calculations on the uploaded data
def run_calculation():
    global output_data
    show_loading("Calculating...")

    try:
        numeric_columns = ['hmnh', 'drs', 'tds', 'net drs amt', 'hmnh percentage', 'drs percentage', 'tds', 'net amount']
        
        # Convert all necessary columns to numeric
        for col in numeric_columns:
            data[col] = pd.to_numeric(data[col], errors='coerce').fillna(0)

        output_data = []
        for _, row in data.iterrows():
            doctor_name = row['performing doctor name']
            sys_net_amt = row['net amount']
            hmnh = row['hmnh'] * row['hmnh percentage'] / 100
            drs = row['drs'] * row['drs percentage'] / 100
            tds_amt = row['tds'] * row['tds'] / 100
            net_drs_amt = row['net drs amt'] * row['drs percentage'] / 100
            
            output_data.append([doctor_name, sys_net_amt, hmnh, drs, tds_amt, net_drs_amt])

        for row in treeview.get_children():
            treeview.delete(row)

        result_columns = ["Doctor Name", "Net Amount", "HMNH", "DRS", "TDS", "Net DRS Amount"]
        treeview["columns"] = result_columns
        treeview["show"] = "headings"

        for col in result_columns:
            treeview.heading(col, text=col, anchor="w")
            treeview.column(col, width=150, anchor="w")

        for row in output_data:
            treeview.insert("", "end", values=row)

        loading_label.config(text="Calculation completed successfully.")
        download_button.pack(side=tk.LEFT, padx=10)
        download_unique_button.pack(side=tk.LEFT, padx=10)


    finally:
        hide_loading()

# Function to start the download process
def start_download():
    threading.Thread(target=download_excel).start()

# Function to download the output data as an Excel file with cell borders
def download_excel():
    show_loading("Downloading Excel...")
    
    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Save as")
        if file_path:
            wb = Workbook()
            ws = wb.active
            ws.append(["Doctor Name", "Net Amount", "HMNH", "DRS", "TDS", "Net DRS Amount"])
            
            for row in output_data:
                ws.append(row)
            
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
            
            wb.save(file_path)
            messagebox.showinfo("Success", "File saved successfully!")
    finally:
        hide_loading()


def start_download_unique():
    threading.Thread(target=download_unique_excel).start()

# Function to download unique calculation results
def download_unique_excel():
    show_loading("Downloading Unique Calculation Excel...")

    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Save as")

        if file_path:
            # Create DataFrame with correct columns
            unique_df = pd.DataFrame(output_data, columns=["Doctor Name", "Net Amount", "HMNH", "DRS", "TDS", "Net DRS Amount"])

            # Convert Doctor Name to string, strip spaces, and remove empty/invalid rows
            unique_df["Doctor Name"] = unique_df["Doctor Name"].astype(str).str.strip()

            # Remove rows where "Doctor Name" is empty, NaN, or "--"
            unique_df = unique_df[~unique_df["Doctor Name"].isin(["", "nan", "NaN", "--"])]

            # Perform groupby and sum
            unique_df = unique_df.groupby("Doctor Name", as_index=False).sum()

            # Debugging: Print cleaned data
            print("Final Data Preview Before Writing to Excel:")
            print(unique_df.head(5))

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active

            # Add headers from DataFrame
            ws.append(list(unique_df.columns))

            # Append cleaned data
            for row in unique_df.itertuples(index=False, name=None):
                ws.append(row)

            # Apply border styling to all cells
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border

            # Save the Excel file
            wb.save(file_path)
            messagebox.showinfo("Success", "Unique Calculation File saved successfully!")

    finally:
        hide_loading()

        
# Start the main event loop
root.mainloop()
